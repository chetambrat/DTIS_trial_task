from datetime import datetime

import openpyxl
import pandas as pd
from selenium import webdriver

path_to_gecko = r".\geckodriver.exe"


def parse_cities(cities_df):
    driver = webdriver.Firefox(executable_path=path_to_gecko)

    cities_dict = {}
    for city in cities_df['City']:
        city_celsius_temp = get_temprerature(driver, city, "celsius")
        city_fahrenheit_temp = get_temprerature(driver, city, "fahrenheit")
        cities_dict.update({city: [city_celsius_temp, city_fahrenheit_temp]})

    driver.close()
    return cities_dict


def get_temprerature(selenium_driver, city, degrees):
    selenium_driver.get(f"https://www.google.com/search?q={city}+weather+{degrees}")
    temperature_element = selenium_driver.find_element_by_id("wob_tm")
    return temperature_element.text


def store_temperatures(cities_dict):
    now = datetime.now()
    current_date = now.strftime("%d.%m.%Y")
    current_time = now.strftime("%H.%M.%S")

    wb = openpyxl.load_workbook('cities.xlsx')
    list_of_worksheets = wb.sheetnames
    ws = wb[list_of_worksheets[0]]
    for row, (key, temps) in enumerate(cities_dict.items(), start=2):
        ws[f"A{row}"] = key
        ws[f"B{row}"] = temps[0]
        ws[f"C{row}"] = temps[1]

    wb.save(f"cities_{current_date}_{current_time}.xlsx")


if __name__ == "__main__":
    cities_names = pd.read_excel('cities.xlsx', index_col=None, header=0)
    cities_stored = parse_cities(cities_names)
    store_temperatures(cities_stored)
